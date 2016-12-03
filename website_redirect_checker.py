#!/usr/bin/python

from selenium import webdriver
from openpyxl import Workbook
import datetime
import random
import urllib2
import time




urls=[

'http://www.centralatlantic.buyatoyota.com/en/specials/camry/',
'http://www.centralatlantic.buyatoyota.com/en/specials/corolla',
'http://www.centralatlantic.buyatoyota.com/en/specials/rav4',
'http://www.centralatlantic.buyatoyota.com/en/specials/highlander',
'http://www.centralatlantic.buyatoyota.com/en/specials/sienna',
'http://www.centralatlantic.buyatoyota.com/en/specials/4runner',
'http://www.centralatlantic.buyatoyota.com/en/specials/tundra',
'http://www.centralatlantic.buyatoyota.com/en/specials/tacoma',
'http://www.centralatlantic.buyatoyota.com/en/specials/prius',

]



def launch_driver(url):
    #make driver global
    global driver

    driver = webdriver.PhantomJS(service_args=['--ssl-protocol=any'])
    driver.maximize_window()

    driver.get(url)
    return



def run_check():
    load=False
    while load==False:
        try:
            current=driver.current_url
            print current
            if 'lease' in str(current):
                global status
                status='Lease Redirect'
                return
            else:
                global status
                status='No Redirect'
                return
        except httplib.BadStatusLine:
            time.sleep(5)
            print ('http error reading...')


def write_to_excel():

    ws['A1'] = 'Model:'
    ws['A2'] = 'Camry'
    ws['A3'] = 'Corolla'
    ws['A4'] = 'Rav4'
    ws['A5'] = 'Highlander'
    ws['A6'] = 'Sienna'
    ws['A7'] = '4Runner'
    ws['A8'] = 'Tundra'
    ws['A9'] = 'Tacoma'
    ws['A10'] = 'Prius'
    ws['B1'] = 'Offer Listing Status:'
    ws['B'+str(i)] = status

def send_message():
    import smtplib
    from email.MIMEMultipart import MIMEMultipart
    from email.MIMEText import MIMEText
    from email.MIMEBase import MIMEBase
    from email import encoders

    fromaddr = '<from email address>'
    toaddr = '<to email address>'

    msg = MIMEMultipart()

    msg['From'] = fromaddr
    msg['To'] = toaddr
    msg['Subject'] = "Toyota Ad Update"

    body = "The status of the Toyota offer listings has changed. Please refer to the attached file for the current offer listing status."

    msg.attach(MIMEText(body, 'plain'))

    filename = "Offer_Listing_Check_new.xlsx"
    attachment = open("Offer_Listing_Check_new.xlsx", "rb")

    part = MIMEBase('application', 'octet-stream')
    part.set_payload((attachment).read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment',filename='Offer_Listing_Check_new.xlsx')

    msg.attach(part)

    server = smtplib.SMTP('smtp.<DOMAIN_NAME>', 587)
    server.starttls()
    server.login(fromaddr, '<password>')
    text = msg.as_string()
    server.sendmail(fromaddr, toaddr, text)
    server.quit()

def read_from_excel():
    from openpyxl import load_workbook
    import openpyxl
    wb = load_workbook(filename = 'Offer_Listing_Check_old.xlsx')

    status_list_old=[]
    status_list.append(wb['Sheet']['B2'].value)
    status_list.append(wb['Sheet']['B3'].value)
    status_list.append(wb['Sheet']['B4'].value)
    status_list.append(wb['Sheet']['B5'].value)
    status_list.append(wb['Sheet']['B6'].value)
    status_list.append(wb['Sheet']['B7'].value)
    status_list.append(wb['Sheet']['B8'].value)
    status_list.append(wb['Sheet']['B9'].value)
    status_list.append(wb['Sheet']['B10'].value)

    wb = load_workbook(filename = 'Offer_Listing_Check_new.xlsx')

    status_list_new=[]
    status_list2.append(wb['Sheet']['B2'].value)
    status_list2.append(wb['Sheet']['B3'].value)
    status_list2.append(wb['Sheet']['B4'].value)
    status_list2.append(wb['Sheet']['B5'].value)
    status_list2.append(wb['Sheet']['B6'].value)
    status_list2.append(wb['Sheet']['B7'].value)
    status_list2.append(wb['Sheet']['B8'].value)
    status_list2.append(wb['Sheet']['B9'].value)
    status_list2.append(wb['Sheet']['B10'].value)
    #print status_list2
    if status_list_old == status_list_new:
        print 'Nothing has changed!'
    else:
        print 'They are different!'
        wb.save('Offer_Listing_Check_old.xlsx')
        send_message()

wb = Workbook()
#grab the active worksheet
ws = wb.active
    #wb.save('Offer_Listing_Check.xlsx')

def check_ads():
    global i
    i=2
    for link in urls:
        launch_driver(link)
        run_check()
        write_to_excel()
        driver.quit()
        i+=1
    wb.save('Offer_Listing_Check_new.xlsx')

check_ads()
read_from_excel()
